import pandas as pd
import openpyxl
from openpyxl import load_workbook
#đọc file
doc = pd.read_excel(r"C:\Users\ACER\Downloads\Temperature.xlsx")
print(doc)


#tải file excel
workbook = load_workbook(filename =r"C:\Users\ACER\Downloads\Temperature.xlsx")
#chọn sheet mặc định
sheet = workbook.active
#lấy giá trị của ô A1
def lay_gia_tri_ô():
    o = input('nhập ô: ')
    value = sheet[o].value
    print("giá trị ô :",o,"là ",value)
# #lấy giá trị của ô bất kì
# def lay_gt_o_bat_ki():
#     print("lấy giá trị của ô bất kì")
#     row = int(input('nhập số hàng: '))
#     col = int(input('nhập số cột: '))
#     value2 = sheet.cell(row=row, column=col).value
#     print(value2)


#đọc nhiều ô cùng một lúc bằng cách duyệt qua nhiều ô
def doc_nhieu_o():
    print('đọc nhiều ô')
    h1=int(input('hàng bắt đầu lấy giá trị: '))
    hn=int(input('đến hàng: '))
    c1=int(input('cột bắt đầu lấy gía trị: '))
    cn=int(input('đến cột:'))
    print("lấy các ô từ hàng ",h1, "đến hàng ",hn,"cột ",c1,"đến cột ",cn)
    for row in sheet.iter_rows(min_row=h1, min_col=hn, max_row=c1, max_col=cn, values_only=True):
        for cell in row:
            print(cell)

#dùng iter_row để duyệt các ô hàng 1 đén 10, cột 1 đến 3
#value_only để lây giá trị của từng ô



#SẮP XẾP THEO THUỘC TÍNH
#sắp xếp theo tăng dần của cột Temperature
def sap_xep_tang():
    ten=input('nhập tên cột cần sắp xếp tăng: ')
    doc_sorted = doc.sort_values(by=[ten])
    print('dữ liệu đã sắp xếp')
    print(doc_sorted)


#hoặc sắp xếp theo thứ tự giảm dần
def sap_xep_giam():
    ten=input("sắp xếp theo thứ tự giảm côt: ")
    doc_sorted2 = doc.sort_values(by=[ten], ascending=False)
    print("du lieu sau khi sắp xếp")
    print(doc_sorted2)


#tham số ascending để sắp xếp theo thứ tự giảm dần

#SẮP XẾP NHIỀU THUỘC TÍNH: Temperature theo thứ tự giảm dần và Salinity theo thứ tự tăng dần
def sap_xep_nhieu_thuoctinh():
    thuoctinh1 = input('nhập tên thuộc tính: ')
    thuoctinh2 = input('nhập tên thuộc tính: ')
    doc_sorte3 = doc.sort_values(by=[thuoctinh1,thuoctinh2],ascending=[False,True])
    print('Dữ liệu sau khi sắp xếp')
    print(doc_sorte3)

#lấy danh sách cách giá trị trong cột Season
#chưa lọc giá trị, danh sách in ra sẽ bị trùng lặp
def lay_cac_gia_tri_cot():
    cot=input("nhập tên cột cần lấy tất cả các giá trị: ")
    names = doc[cot].tolist()  #tolist chuyển đổi 1 cột của object thành 1 list trong python
    print("in ra danh sách các giá trị trong cột",cot,names)

#lọc các giá trị và in ra danh sách các giá trị không bị trùng lặp
def lay_gt_cot_khong_lap():
    cot=input('nhập tên cột lây GT không lặp,(vd:Season): ')
    loc = doc[cot].unique()
    for value in loc:
        print(value)

#tính max, min, tổng, trung bình,...
def max_min():
    cot=input('nhap tên cột cần tìm max, min: ')
    max = doc[cot].max()
    min = doc[cot].min()
    print("giá trị lớn nhất:",max, "giá trị nhỏ nhất:",min)
def tong_mean():
    cot = input('nhap tên cột cần tính tổng và trung bình: ')
    tong = doc['Temperature'].sum()
    print("Tổng là:",tong)
    trung_binh = doc[cot].mean()
    print("giá trị trung bình của cột Temperature là: ",trung_binh)

def main():
    # đọc file
    doc = pd.read_excel(r"C:\Users\ACER\Downloads\Temperature.xlsx")
    print(doc)
    lay_gt_cot_khong_lap()
    lay_cac_gia_tri_cot()
    lay_gia_tri_ô()
    # lay_gt_o_bat_ki()
    doc_nhieu_o()
    sap_xep_giam()
    sap_xep_tang()
    sap_xep_nhieu_thuoctinh()
    max_min()
    tong_mean()

if __name__ == '__main__':
    main()

